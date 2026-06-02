"""Importador del Libro de Caja desde xlsx.

El libro de caja tiene una zona de "datos" con 3-5 columnas:
    1: Fecha (puede estar vacía en filas del mismo día; se hereda de la anterior)
    2: Denominación
    3: Importe (positivo=ingreso, negativo=gasto)
    4: Saldo acumulado (no se importa, sólo verificación en Excel)
    5: Observaciones

La zona de "instrucciones" al final (con cabeceras "SI EL IMPORTE ES POSITIVO
ENTONCES:") debe ignorarse.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl


SECCION_BLOQUEO = {
    "si el importe del asiento es positivo entonces:",
    "si el importe del asiento es negativo entonces:",
    "si el importe del asiento es",
    "si movimiento=scf-traspaso fondos e importe de asiento positivo",
    "si movimiento=scf-traspaso fondos e importe de asiento negativo",
    "los comentarios se añadirian manualmente en el programa",
}


@dataclass(frozen=True)
class LineaCaja:
    fecha: date
    denominacion: str
    importe: float
    observaciones: str = ""

    @property
    def signo(self) -> int:
        return 1 if self.importe >= 0 else -1


def _a_fecha(v) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        try:
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=float(v))).date()
        except (OverflowError, ValueError):
            return None
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _a_float(v) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("€", "").replace(" ", "").replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _es_numero_sospechoso(original, fecha_parseada: date) -> bool:
    """Enteros pequeños que parecen "1, 2, 3" del Excel no son fechas reales."""
    if isinstance(original, (int, float)) and not isinstance(original, bool):
        if 1 <= float(original) <= 31:
            return True
    return False


def _es_bloque_seccion(texto: str) -> bool:
    t = texto.strip().lower()
    if not t:
        return False
    if t in {"d", "h", "ordenmovimiento", "cargoabono", "codigocuenta", "si", "entonces"}:
        return True
    for clave in SECCION_BLOQUEO:
        if clave in t:
            return True
    return False


def _detectar_cabecera_caja(ws) -> int | None:
    """Busca fila con cabeceras 'Fecha del asiento' / 'comentario' / 'IMPORTE'."""
    for r in range(1, min(ws.max_row, 30) + 1):
        row = [str(ws.cell(r, c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
        joined = " | ".join(row)
        if "fecha del asiento" in joined and "importe" in joined:
            return r
    return None


def _buscar_hoja_caja(wb) -> object | None:
    for name in wb.sheetnames:
        ws = wb[name]
        if _detectar_cabecera_caja(ws) is not None:
            return ws
    return None


def importar_libro_caja(path: str | Path) -> list[LineaCaja]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = _buscar_hoja_caja(wb) or wb.active
    cab = _detectar_cabecera_caja(ws)
    if cab is None:
        return []
    start = cab + 1

    lineas: list[LineaCaja] = []
    # En el excel del usuario la fecha real NO está: se rellena en la UI
    fecha_vigente: date | None = None

    for r in range(start, ws.max_row + 1):
        v1 = ws.cell(r, 1).value
        v2 = ws.cell(r, 2).value
        v3 = ws.cell(r, 3).value
        v5 = ws.cell(r, 5).value

        # Detección de sección explicativa
        if v1 is not None and _es_bloque_seccion(str(v1)):
            break
        if v2 is not None and _es_bloque_seccion(str(v2)):
            break

        # Línea de saldo inicial explícito: ignorar (no genera asiento)
        denom_low = "" if v2 is None else str(v2).strip().lower()
        if denom_low == "saldo inicial":
            continue

        denom_raw = v2
        denom = "" if denom_raw is None else str(denom_raw).strip()
        importe = _a_float(v3)

        # Fila vacía real: nada de nada
        if not denom and importe is None:
            continue
        if not denom or importe is None:
            continue

        # Si parece una fecha real, la guardamos; si no, queda None
        fecha_candidata = _a_fecha(v1)
        if fecha_candidata is not None and not _es_numero_sospechoso(v1, fecha_candidata):
            fecha_vigente = fecha_candidata

        obs = "" if v5 is None else str(v5).strip()
        lineas.append(LineaCaja(fecha_vigente, denom, float(importe), obs))

    return lineas
