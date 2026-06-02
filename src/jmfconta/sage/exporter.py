"""Exportador a xlsx con el formato esperado por SAGE para importar asientos.

Columnas (en este orden exacto, fila 1 = cabecera):
    Asiento | Numerodeperiodo | OrdenMovimiento | CargoAbono |
    CodigoCuenta | FechaAsiento | ImporteAsiento | Comentario
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

from .rules import AsientoGenerado

CABECERA = [
    "Asiento",
    "Numerodeperiodo",
    "OrdenMovimiento",
    "CargoAbono",
    "CodigoCuenta",
    "FechaAsiento",
    "ImporteAsiento",
    "Comentario",
]


def exportar_sage(asientos: list[AsientoGenerado], path: str | Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IMPORTAR A SAGE"

    # Cabecera en negrita
    for col, val in enumerate(CABECERA, start=1):
        c = ws.cell(1, col, val)
        c.font = Font(bold=True)

    row = 2
    for n_asiento, ast in enumerate(asientos, start=1):
        for linea in ast.lineas:
            ws.cell(row, 1, n_asiento)
            ws.cell(row, 2, ast.periodo)
            ws.cell(row, 3, linea.orden)
            ws.cell(row, 4, linea.cargo_abono)
            ws.cell(row, 5, linea.cuenta)
            ws.cell(row, 6, ast.fecha)
            ws.cell(row, 6).number_format = "yyyy-mm-dd"
            ws.cell(row, 7, float(linea.importe))
            ws.cell(row, 7).number_format = "0.00"
            ws.cell(row, 8, linea.comentario or ast.descripcion)
            row += 1

    # Anchos de columna razonables
    anchos = [9, 17, 16, 12, 13, 13, 15, 40]
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(path)
