"""Smoke test headless: instancia la app, carga excels reales, ejercita cada pestaña."""
import os
import sys
import tempfile
from pathlib import Path

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")

from PySide6.QtWidgets import QApplication  # noqa: E402

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from jmfconta import db, repository  # noqa: E402
from jmfconta.importers.banco import importar_movimientos_banco  # noqa: E402
from jmfconta.importers.caja import importar_libro_caja  # noqa: E402
from jmfconta.importers.plan_cuentas import importar_plan_cuentas  # noqa: E402
from jmfconta.sage.exporter import exportar_sage  # noqa: E402
from jmfconta.ui.main_window import MainWindow  # noqa: E402

DOCS = Path(__file__).resolve().parent.parent / "docs"


def run():
    with tempfile.TemporaryDirectory() as tmp:
        db_path = Path(tmp) / "t.db"
        db.init_db(db_path)
        conn = db.connect(db_path)

        # Cargar plan
        n = repository.cargar_plan(conn, importar_plan_cuentas(DOCS / "PLAN DE CUENTAS JM.xlsx"))
        print(f"[OK] plan cargado: {n} cuentas")

        # Cargar caja
        lineas_caja = importar_libro_caja(DOCS / "Ejemplo subida SAGE.xlsx")
        n = repository.insertar_movimientos_caja(conn, lineas_caja)
        print(f"[OK] caja importada: {n} líneas (negativas: {sum(1 for l in lineas_caja if l.importe < 0)})")

        # Simular que el usuario rellena fechas y asigna cuentas a la caja
        for row in repository.listar_movimientos_caja(conn):
            # Fecha por defecto: 2026-05-14 (la del ejemplo SAGE) si está vacía
            if not row["fecha"]:
                repository.actualizar_fecha_caja(conn, row["id"], "2026-05-14")
            # Cuenta por defecto: 7053001 ingresos / 6280001 gastos
            if not row["cuenta_sugerida"]:
                cuenta = "7053001" if row["importe"] > 0 else "6280001"
                repository.actualizar_cuenta_caja(conn, row["id"], cuenta)

        # Cargar banco
        lineas_banco = importar_movimientos_banco(DOCS / "Movimientos_cuenta_banco.xls")
        n = repository.insertar_movimientos_banco(conn, lineas_banco)
        print(f"[OK] banco importado: {n} movimientos (barridos: {sum(1 for l in lineas_banco if l.es_barrido)})")

        # Asignar mappings: para los movimientos de banco, usar el campo 'mas_datos' como clave
        # (los movimientos ya se autocompletaron si había mapping previo)
        for row in repository.listar_movimientos_banco(conn):
            if not row["cuenta_sugerida"]:
                # Por defecto, ingresos al 7053001, gastos al 6280001
                cuenta = "7053001" if row["importe"] > 0 else "6280001"
                repository.actualizar_cuenta_banco(conn, row["id"], cuenta)
                if row["mas_datos"]:
                    repository.set_mapping(conn, "BANCO", row["mas_datos"], cuenta)
        for row in repository.listar_movimientos_caja(conn):
            if not row["cuenta_sugerida"]:
                cuenta = "7053001" if row["importe"] > 0 else "6280001"
                repository.actualizar_cuenta_caja(conn, row["id"], cuenta)
                repository.set_mapping(conn, "CAJA", row["denominacion"], cuenta)
        print("[OK] mappings asignados por defecto")

        # Generar pre-asientos
        asientos = repository.generar_asientos_caja(conn) + repository.generar_asientos_banco(conn)
        print(f"[OK] asientos generados: {len(asientos)}")
        assert len(asientos) > 0

        # Exportar SAGE
        out = Path(tmp) / "asientos_sage.xlsx"
        exportar_sage(asientos, out)
        assert out.exists()
        print(f"[OK] exportado: {out} ({out.stat().st_size} bytes)")

        # Instanciar UI (offscreen)
        app = QApplication.instance() or QApplication(sys.argv)
        win = MainWindow(conn)
        win.show()
        # Refrescar y verificar que las tablas no explotan
        for tab in (win.tabs.widget(i) for i in range(win.tabs.count())):
            if hasattr(tab, "_refill"):
                try:
                    tab._refill()
                except TypeError:
                    tab._refill("")
        print("[OK] UI instanciada y refrescos sin error")
        # Imprimir las primeras 8 líneas del export para verificación humana
        from openpyxl import load_workbook
        wb = load_workbook(out, data_only=True)
        ws = wb.active
        print("\n--- SAGE export preview ---")
        for r in range(1, min(ws.max_row + 1, 20)):
            print(" | ".join(str(ws.cell(r, c).value or "") for c in range(1, 9)))


if __name__ == "__main__":
    run()
