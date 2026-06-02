from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from jmfconta.sage.exporter import exportar_sage
from jmfconta.sage.rules import generar_desde_banco, generar_desde_caja


def test_export_formato_correcto(tmp_path: Path):
    asientos = [
        generar_desde_caja(date(2026, 5, 14), 100.0, "7053001", "SERV. COMPLE"),
        generar_desde_caja(date(2026, 5, 14), -24.0, "6280001", "GASOLINA"),
        generar_desde_banco(date(2026, 5, 18), 60.0, "TRANSFER INMEDIATA", "4300001", "LAURA"),
        generar_desde_banco(date(2026, 5, 18), 60.0, "SCF-TRASPASO FONDOS", "", "TRASPASO"),
    ]
    out = tmp_path / "sage.xlsx"
    exportar_sage(asientos, out)

    wb = load_workbook(out, data_only=True)
    ws = wb.active
    cab = [ws.cell(1, c).value for c in range(1, 9)]
    assert cab == [
        "Asiento", "Numerodeperiodo", "OrdenMovimiento", "CargoAbono",
        "CodigoCuenta", "FechaAsiento", "ImporteAsiento", "Comentario",
    ]
    # 4 asientos x 2 lineas = 8 filas de datos
    assert ws.max_row == 9
    # asiento 3 linea 2: H 4300001
    fila = None
    for r in range(2, 10):
        if ws.cell(r, 1).value == 3 and ws.cell(r, 3).value == 2:
            fila = r
            break
    assert fila is not None
    assert ws.cell(fila, 4).value == "H"
    assert ws.cell(fila, 5).value == "4300001"
    assert ws.cell(fila, 7).value == 60.0
